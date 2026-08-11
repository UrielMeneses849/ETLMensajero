# -*- coding: utf-8 -*-

import json
import os
import sys
from datetime import datetime
from io import BytesIO
from typing import Optional

import pandas as pd

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU


# =========================================================
# CONFIG
# =========================================================

AUTO_ROW_HEIGHT = False

AUTO_ROW_MIN_HEIGHT = 34
AUTO_ROW_MAX_HEIGHT = 260
AUTO_LINE_HEIGHT = 23

EXCEL_WEB_MODE = True
EXCEL_WEB_MIN_ROW_HEIGHT = 38

COLOR_ORANGE = "ED7D31"
COLOR_GRAY = "E7E6E6"


# =========================================================
# HELPERS DE ENCABEZADOS
# =========================================================

def _export_headers_with_spaces(cols):
    """
    ÚNICAMENTE modifica cómo se muestran los encabezados en Excel.

    Ejemplo:
        orden_de_gobierno
    se muestra como:
        orden de gobierno

    Esto NO toca los valores de la data.
    """

    return [
        str(col).replace("_", " ").strip()
        for col in cols
    ]


# =========================================================
# ANCHOS DE COLUMNA
# =========================================================

def _compute_widths_from_df(
    df: pd.DataFrame,
    padding: int = 4,
    max_width: int = 60
) -> dict:

    if df.empty:
        return {}

    # Para no recorrer datasets gigantes completos
    sample = df if len(df) <= 2000 else df.head(2000)

    widths = {}

    for col in sample.columns:

        max_len = len(str(col))

        for value in sample[col]:

            if value is None:
                continue

            value_length = len(str(value))

            if value_length > max_len:
                max_len = value_length

        widths[col] = min(
            max_len + padding,
            max_width
        )

    return widths


def _normalize_column_name_for_style(name: str) -> str:
    """
    Normaliza únicamente el nombre de columna para decidir
    estilos visuales.

    NO modifica la data.
    """

    if name is None:
        return ""

    return (
        str(name)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )


def _apply_width_overrides(
    ws,
    df_export: pd.DataFrame
):

    """
    Anchos especiales únicamente de presentación.
    """

    TARGET_WIDTHS = {

        "email_1": 42,
        "email_2": 42,
        "email_3": 55,

        "proyecto": 95,

        "descripcion": 120,
        "descripcionextra": 120,
        "descripcion_extra": 120,

        "observaciones": 95,

        "localizacion": 85,
        "localizacion1": 85,
        "localizacion_del_proyecto": 95,

        "acabados": 120,
    }

    for idx, col_name in enumerate(
        df_export.columns,
        start=1
    ):

        key = _normalize_column_name_for_style(
            col_name
        )

        if key == "email1":
            key = "email_1"

        elif key == "email2":
            key = "email_2"

        elif key == "email3":
            key = "email_3"

        elif key == "descripcion_extra_del_proyecto":
            key = "descripcion_extra"

        if key in TARGET_WIDTHS:

            ws.column_dimensions[
                get_column_letter(idx)
            ].width = TARGET_WIDTHS[key]


# =========================================================
# RECURSOS
# =========================================================

def _resolve_resource_path(
    filename: str
) -> Optional[str]:

    """
    Busca recursos como logo_bimsa.jpg
    tanto en PyInstaller como ejecución normal.
    """

    base_path = getattr(
        sys,
        "_MEIPASS",
        None
    )

    if base_path:

        candidate = os.path.join(
            base_path,
            filename
        )

        if os.path.exists(candidate):
            return candidate

    candidate = os.path.join(
        os.path.dirname(__file__),
        filename
    )

    if os.path.exists(candidate):
        return candidate

    candidate = os.path.join(
        os.getcwd(),
        filename
    )

    if os.path.exists(candidate):
        return candidate

    return None


# =========================================================
# ESTILOS DEL EXCEL
# =========================================================

def _apply_styles_excel_and_sheets(
    ws,
    header_row: int,
    first_data_row: int,
    nrows: int,
    ncols: int
):

    """
    Aplica solamente estilos visuales.

    IMPORTANTE:
    Esta función NO modifica cell.value.
    """

    header_font = Font(
        name="Poppins",
        size=11,
        bold=True,
        color="FFFFFF"
    )

    body_font = Font(
        name="Poppins",
        size=11
    )

    header_fill = PatternFill(
        "solid",
        COLOR_ORANGE
    )

    zebra_a = PatternFill(
        "solid",
        "F2F2F2"
    )

    zebra_b = PatternFill(
        "solid",
        "FFFFFF"
    )

    header_align = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    body_align_center = Alignment(
        wrap_text=True,
        vertical="center",
        horizontal="center"
    )

    body_align_left = Alignment(
        wrap_text=True,
        vertical="center",
        horizontal="left"
    )

    # Solamente afecta alineación.
    LEFT_ALIGN_COLUMNS = {

        "proyecto",
        "localizacion1",
        "localizacion",

        "descripcion",
        "descripcion_del_proyecto",

        "acabados",
        "observaciones",
    }

    # Congelar encabezados
    ws.freeze_panes = f"A{first_data_row}"

    # Ocultar cuadrícula
    ws.sheet_view.showGridLines = False

    # Filtros
    if ncols > 0 and nrows >= header_row:

        ws.auto_filter.ref = (
            f"A{header_row}:"
            f"{get_column_letter(ncols)}{nrows}"
        )

    # =====================================================
    # HEADER
    # =====================================================

    for cell in ws[header_row]:

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = header_align

    # =====================================================
    # BODY
    # =====================================================

    for row_idx in range(
        first_data_row,
        nrows + 1
    ):

        fill = (
            zebra_a
            if row_idx % 2 == 0
            else zebra_b
        )

        for col_idx in range(
            1,
            ncols + 1
        ):

            cell = ws.cell(
                row=row_idx,
                column=col_idx
            )

            header_name = str(
                ws.cell(
                    row=header_row,
                    column=col_idx
                ).value
            )

            normalized_header = (
                header_name
                .lower()
                .replace(" ", "_")
            )

            cell.font = body_font

            cell.fill = fill

            if normalized_header in LEFT_ALIGN_COLUMNS:

                cell.alignment = body_align_left

            else:

                cell.alignment = body_align_center


# =========================================================
# ALTURA FIJA
# =========================================================

def _apply_fixed_row_height(
    ws,
    first_data_row: int,
    nrows: int,
    height: int = 60
):

    for row_idx in range(
        first_data_row,
        nrows + 1
    ):

        ws.row_dimensions[
            row_idx
        ].height = height


# =========================================================
# AUTO HEIGHT
# =========================================================

def _safe_col_width(
    ws,
    col_letter: str
) -> float:

    width = ws.column_dimensions[
        col_letter
    ].width

    if width is not None and width > 0:
        return float(width)

    return 12.0


def _estimate_wrapped_lines(
    text,
    col_width_chars: float
) -> int:

    if text is None:
        return 1

    text = str(text)

    if not text:
        return 1

    SHEETS_WIDTH_FACTOR = 0.88

    effective_width = max(
        8,
        int(
            col_width_chars
            * SHEETS_WIDTH_FACTOR
        ) - 1
    )

    total_lines = 0

    for raw_line in text.split("\n"):

        raw_line = raw_line.strip()

        if not raw_line:

            total_lines += 1

            continue

        words = raw_line.split()

        current_length = 0

        lines = 1

        for word in words:

            word_length = len(word)

            if current_length == 0:

                current_length = word_length

            elif (
                current_length
                + 1
                + word_length
                <= effective_width
            ):

                current_length += (
                    1
                    + word_length
                )

            else:

                lines += 1

                current_length = word_length

        total_lines += lines

    return max(
        1,
        total_lines
    )


def _apply_auto_row_heights(
    ws,
    first_data_row: int,
    min_h: int,
    max_h: int,
    line_h: int
):

    """
    Calcula alto de filas.

    Solo usa cell.value para medir texto.

    NO modifica el contenido.
    """

    for row_idx in range(
        first_data_row,
        ws.max_row + 1
    ):

        max_lines = 1

        for col_idx in range(
            1,
            ws.max_column + 1
        ):

            cell = ws.cell(
                row=row_idx,
                column=col_idx
            )

            if cell.value is None:
                continue

            col_letter = get_column_letter(
                col_idx
            )

            col_width = _safe_col_width(
                ws,
                col_letter
            )

            lines = _estimate_wrapped_lines(
                cell.value,
                col_width
            )

            max_lines = max(
                max_lines,
                lines
            )

        calculated_height = int(
            max_lines
            * line_h
        )

        final_height = max(
            min_h,
            min(
                max_h,
                calculated_height
            )
        )

        ws.row_dimensions[
            row_idx
        ].height = final_height


# =========================================================
# BRANDING BIMSA
# =========================================================

def _apply_branding_row(
    ws,
    ncols: int,
    empresa: str,
    usuario: str,
    report_label: str,
    logo_filename: str = "logo_bimsa.jpg",
    logo_path: Optional[str] = None
):

    """
    Crea primera fila con branding BIMSA.

    Esto no toca valores provenientes del backend.
    """

    ws.row_dimensions[1].height = 50

    ws.column_dimensions[
        "A"
    ].width = 28

    white_fill = PatternFill(
        "solid",
        "FFFFFF"
    )

    bottom_side = Side(
        style="thin",
        color="000000"
    )

    # =====================================================
    # FONDO / BORDE
    # =====================================================

    for col_idx in range(
        1,
        ncols + 1
    ):

        cell = ws.cell(
            row=1,
            column=col_idx
        )

        cell.fill = white_fill

        existing_border = cell.border

        cell.border = Border(
            left=existing_border.left,
            right=existing_border.right,
            top=existing_border.top,
            bottom=bottom_side
        )

    # =====================================================
    # LOGO
    # =====================================================

    if (
        logo_path
        and os.path.exists(logo_path)
    ):

        img = XLImage(
            logo_path
        )

        img.width = 170

        img.height = 40

        cell_height_emu = pixels_to_EMU(
            55
        )

        img_height_emu = pixels_to_EMU(
            img.height
        )

        vertical_offset = int(
            (
                cell_height_emu
                - img_height_emu
            )
            / 2
        )

        horizontal_offset = pixels_to_EMU(
            10
        )

        marker = AnchorMarker(
            colOff=horizontal_offset,
            rowOff=vertical_offset
        )

        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(
                pixels_to_EMU(
                    img.width
                ),
                pixels_to_EMU(
                    img.height
                )
            )
        )

        ws.add_image(
            img
        )

    # =====================================================
    # EMPRESA / USUARIO
    # =====================================================

    if ncols >= 4:

        ws.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=4
        )

    elif ncols >= 2:

        ws.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=ncols
        )

    info_cell = ws.cell(
        row=1,
        column=2
    )

    info_cell.value = (
        f"Empresa: {empresa}\n"
        f"Usuario: {usuario}"
    )

    info_cell.font = Font(
        name="Poppins",
        size=11,
        bold=False,
        color="000000"
    )

    info_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    # =====================================================
    # TIPO REPORTE
    # =====================================================

    if ncols >= 2:

        tag_start_col = max(
            1,
            ncols - 1
        )

        # Evitar conflicto con merge B:D
        if tag_start_col > 4:

            ws.merge_cells(
                start_row=1,
                start_column=tag_start_col,
                end_row=1,
                end_column=ncols
            )

            tag_cell = ws.cell(
                row=1,
                column=tag_start_col
            )

            tag_cell.value = report_label

            tag_cell.font = Font(
                name="Poppins",
                size=14,
                bold=True,
                color="FFFFFF"
            )

            tag_cell.fill = PatternFill(
                "solid",
                COLOR_ORANGE
            )

            tag_cell.alignment = Alignment(
                horizontal="right",
                vertical="center",
                wrap_text=True
            )


# =========================================================
# NORMALIZACIÓN DE FECHAS
# =========================================================

def _normalize_date_columns(
    df: pd.DataFrame
) -> pd.DataFrame:

    """
    ESTA ES LA ÚNICA TRANSFORMACIÓN DE DATA DEL ETL.

    Toda columna cuyo nombre contiene "fecha"
    se transforma a datetime.

    Ninguna otra columna se toca.
    """

    for col in df.columns:

        column_name = str(
            col
        ).lower()

        if "fecha" not in column_name:
            continue

        df[col] = pd.to_datetime(
            df[col],
            errors="coerce"
        )

    return df


# =========================================================
# FORMATO DE FECHAS EN EXCEL
# =========================================================

def _apply_excel_date_format(
    ws,
    original_columns,
    first_data_row: int
):

    """
    Aplica formato yyyy-mm-dd únicamente
    a columnas identificadas como fecha.

    No modifica otras columnas.
    """

    for col_idx, col_name in enumerate(
        original_columns,
        start=1
    ):

        if (
            "fecha"
            not in str(col_name).lower()
        ):

            continue

        for row_idx in range(
            first_data_row,
            ws.max_row + 1
        ):

            cell = ws.cell(
                row=row_idx,
                column=col_idx
            )

            if isinstance(
                cell.value,
                datetime
            ):

                cell.number_format = (
                    "yyyy-mm-dd"
                )


# =========================================================
# ETL PRINCIPAL
# =========================================================

def ETL_BIMSA(
    ruta_json: str,
    tipo_reporte: str,
    return_mode: str = "file",
    carpeta_excel: str = ".",
    empresa: str = "",
    usuario: str = "",
    tipo_fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    report_label: Optional[str] = None,
    logo_path: Optional[str] = None,
):

    print(
        "[BIMSA_ETL] Iniciando ETL BIMSA..."
    )

    # =====================================================
    # CONFIG REPORTE
    # =====================================================

    tipo_upper = str(
        tipo_reporte
    ).strip().upper()

    now = datetime.now()

    nombre_excel = (
        f"BIMSA_"
        f"{tipo_upper}_"
        f"{now.strftime('%Y%m%d_%H%M%S')}"
        f".xlsx"
    )

    if report_label is None:

        report_label = (
            tipo_upper.lower()
        )

    # =====================================================
    # LEER JSON
    # =====================================================

    with open(
        ruta_json,
        "r",
        encoding="utf-8"
    ) as file:

        data = json.load(
            file
        )

    if (
        not isinstance(data, list)
        or not data
    ):

        raise ValueError(
            "El JSON debe ser una lista "
            "con al menos un registro"
        )

    # =====================================================
    # DATAFRAME
    # =====================================================

    df = pd.DataFrame(
        data
    )

    # =====================================================
    # ÚNICA TRANSFORMACIÓN PERMITIDA
    # =====================================================

    df = _normalize_date_columns(
        df
    )

    # =====================================================
    # COPIA PARA EXPORTACIÓN
    # =====================================================

    df_export = df.copy()

    original_columns = list(
        df.columns
    )

    # =====================================================
    # NULOS
    # =====================================================
    #
    # openpyxl necesita None para celdas vacías.
    #
    # Esto no modifica valores válidos.
    # =====================================================

    df_export = (
        df_export
        .astype(object)
        .where(
            pd.notnull(df_export),
            None
        )
    )

    # =====================================================
    # HEADERS VISUALES
    # =====================================================
    #
    # orden_de_gobierno
    #
    # se mostrará como:
    #
    # orden de gobierno
    #
    # pero la DATA no se modifica.
    # =====================================================

    df_export.columns = (
        _export_headers_with_spaces(
            df_export.columns
        )
    )

    # =====================================================
    # CREAR WORKBOOK
    # =====================================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Reporte"

    header_row = 2

    first_data_row = 3

    # =====================================================
    # ESCRIBIR HEADERS
    # =====================================================

    for col_idx, col_name in enumerate(
        df_export.columns,
        start=1
    ):

        ws.cell(
            row=header_row,
            column=col_idx,
            value=(
                str(col_name)
                if col_name is not None
                else ""
            )
        )

    # =====================================================
    # ESCRIBIR DATA
    # =====================================================
    #
    # IMPORTANTE:
    #
    # Aquí NO:
    #
    # .upper()
    # .lower()
    # .title()
    # .strip()
    # float()
    # int()
    # pd.to_numeric()
    #
    # nada.
    #
    # Se escribe exactamente el valor recibido,
    # excepto las fechas previamente normalizadas.
    # =====================================================

    data_matrix = (
        df_export
        .to_numpy()
    )

    date_column_indexes = set()

    for idx, column_name in enumerate(
        original_columns,
        start=1
    ):

        if (
            "fecha"
            in str(column_name).lower()
        ):

            date_column_indexes.add(
                idx
            )

    for row_offset, row in enumerate(
        data_matrix
    ):

        excel_row = (
            first_data_row
            + row_offset
        )

        for col_idx, value in enumerate(
            row,
            start=1
        ):

            # =============================================
            # ÚNICAMENTE FECHAS
            # =============================================

            if (
                col_idx
                in date_column_indexes
            ):

                if isinstance(
                    value,
                    pd.Timestamp
                ):

                    value = (
                        value
                        .to_pydatetime()
                    )

            # =============================================
            # ESCRIBIR VALOR
            # =============================================

            ws.cell(
                row=excel_row,
                column=col_idx,
                value=value
            )

    # =====================================================
    # DIMENSIONES
    # =====================================================

    ncols = len(
        df_export.columns
    )

    nrows = (
        first_data_row
        + len(df_export)
        - 1
    )

    # =====================================================
    # ANCHOS AUTOMÁTICOS
    # =====================================================

    widths = _compute_widths_from_df(
        df_export,
        padding=4,
        max_width=60
    )

    MIN_HEADER_WIDTH = 18

    for idx, col_name in enumerate(
        df_export.columns,
        start=1
    ):

        header_length = len(
            str(col_name)
        )

        base_width = widths.get(
            col_name,
            12
        )

        header_width = (
            header_length
            + 4
        )

        final_width = max(
            base_width,
            header_width,
            MIN_HEADER_WIDTH
        )

        ws.column_dimensions[
            get_column_letter(idx)
        ].width = final_width

    # =====================================================
    # OVERRIDES VISUALES
    # =====================================================

    _apply_width_overrides(
        ws,
        df_export
    )

    # =====================================================
    # ESTILOS
    # =====================================================

    _apply_styles_excel_and_sheets(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        nrows=nrows,
        ncols=ncols
    )

    # =====================================================
    # BORDE HEADER
    # =====================================================

    header_border = Border(
        bottom=Side(
            style="thin",
            color="000000"
        )
    )

    for col_idx in range(
        1,
        ncols + 1
    ):

        cell = ws.cell(
            row=header_row,
            column=col_idx
        )

        existing_border = (
            cell.border
        )

        cell.border = Border(
            left=existing_border.left,
            right=existing_border.right,
            top=existing_border.top,
            bottom=header_border.bottom
        )

    ws.row_dimensions[
        header_row
    ].height = 38

    # =====================================================
    # BRANDING
    # =====================================================

    if not logo_path:

        logo_path = (
            _resolve_resource_path(
                "logo_bimsa.jpg"
            )
        )

    _apply_branding_row(
        ws,
        ncols=ncols,
        empresa=empresa,
        usuario=usuario,
        report_label=report_label,
        logo_filename="logo_bimsa.jpg",
        logo_path=logo_path
    )

    # =====================================================
    # FORMATO DE FECHAS
    # =====================================================

    _apply_excel_date_format(
        ws,
        original_columns=original_columns,
        first_data_row=first_data_row
    )

    # =====================================================
    # ALTURA DE FILAS
    # =====================================================

    if AUTO_ROW_HEIGHT:

        min_height = (
            AUTO_ROW_MIN_HEIGHT
        )

        if EXCEL_WEB_MODE:

            min_height = max(
                min_height,
                EXCEL_WEB_MIN_ROW_HEIGHT
            )

        _apply_auto_row_heights(
            ws,
            first_data_row=first_data_row,
            min_h=min_height,
            max_h=AUTO_ROW_MAX_HEIGHT,
            line_h=AUTO_LINE_HEIGHT
        )

    else:

        _apply_fixed_row_height(
            ws,
            first_data_row=first_data_row,
            nrows=nrows,
            height=60
        )

    # =====================================================
    # RETURN BYTES
    # =====================================================

    if return_mode == "bytes":

        output = BytesIO()

        wb.save(
            output
        )

        output.seek(
            0
        )

        print(
            "[BIMSA_ETL] "
            "ETL terminado correctamente: "
            f"{nombre_excel}"
        )

        return (
            nombre_excel,
            output.getvalue()
        )

    # =====================================================
    # GUARDAR ARCHIVO
    # =====================================================

    os.makedirs(
        carpeta_excel,
        exist_ok=True
    )

    ruta_final = (
        f"{carpeta_excel.rstrip('/')}/"
        f"{nombre_excel}"
    )

    wb.save(
        ruta_final
    )

    print(
        "[BIMSA_ETL] "
        "ETL terminado correctamente: "
        f"{nombre_excel}"
    )

    return ruta_final